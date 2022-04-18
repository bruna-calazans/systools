# -*- coding: utf-8 -*-
"""
Created on Thu Feb  8 10:55:13 2018

@author: bcalazans
"""
import os
import pandas as pd
import geopandas as gpd

from zipfile import ZipFile
from openpyxl import load_workbook


# HELPERS
def get_american_standers(flag):
    """
    If flag is true, return the american stander decimal and separator.
    :param flag: bool True or False.
    :return: dictionary to read or write csv's.
    """
    if flag:
        keys = {'sep': ',', 'decimal': '.'}
    else:
        keys = {'sep': ';', 'decimal': ','}
    return keys


def file_name_with_extension(path, name, ext='.csv'):
    # Reassures passed name is without extension.
    name = os.path.splitext(os.path.basename(name))[0]
    name_ext = os.path.normpath(os.path.join(path, name)) + ext
    return name_ext


# LOAD FILES
def load_text_file(path, name, cols=None, usa=False, kwargs=None):
    """
    Reads a textfile or excel 
    :param path: path of file, may end with .zip if name inside a zipfile
    :param name: name of file with extension
    :param cols: columns to load
    :param usa: if american is true, decimals=','
    :param kwargs: all kwargs accepted by pd.read_csv as a dictionary
    :return: a dataFrame
    """
    if kwargs is None:
        kwargs = {}

    keywords = get_american_standers(usa)
    kwargs = {**keywords, **kwargs}
    
    if 'engine' not in kwargs.keys(): 
        if 'low_memory' not in kwargs.keys():
            kwargs['low_memory'] = False
    
    # Escolhe entre ler csv ou excel.
    def read(pathname):
        if name.lower().endswith('.xlsx'):            
            kwargs.pop('low_memory', None)
            kwargs.pop('sep', None)
            kwargs.pop('decimal', None)
            df_ltf = pd.read_excel(pathname, usecols=cols, **kwargs)
        else:
            df_ltf = pd.read_csv(pathname, usecols=cols, **kwargs)
        return df_ltf
    
    # Trata caso esteja ZIPADO.
    if path[-3:] == 'zip':
        with ZipFile(os.path.normpath(path)) as z:
            with z.open(name) as myFile:
                df = read(myFile)                
    else:
        df = read(os.path.normpath(os.path.join(path, name)))

    # Drop rows that contains only NaN
    df.dropna(how="all", inplace=True)
    # Drop cols that contains only NaN
    if df.shape[0] > 1:
        df.dropna(how="all", inplace=True, axis=1)

    return df


def load_shp_file(path, name, cols=None):
    """
    Reads a shapefile or a DBF lonly file
    :param path: path of file
    :param name: name of file with extension
    :param cols: columns expected
    :return: return the data as a geodataframe
    """
    df = gpd.read_file(os.path.normpath(os.path.join(path, name)))
    try:
        # Look for the auxiliary document with the compleat name of columns.
        aux = os.path.join(path, name[:-3]+'col')
        aux = pd.read_csv(aux, header=None, engine='python')
        aux_dict = dict(zip(aux[1], aux[0]))
        df = df.rename(columns=aux_dict)
    except FileNotFoundError:
        pass
    
    if cols is not None:
        for c in cols:
            if c not in df.columns:
                msg_error = "File {} is opened somewhere or missing column {}"\
                    .format(name, c)
                raise Exception(msg_error)
            else:
                pass
        df = df.loc[:, cols]  # Return only expected columns
    return df


def load_file(path, name, expected_cols, usa, **kwargs):
    
    if name is None:
        name = os.path.basename(path)
        path = os.path.dirname(path)
    
    # Checks if name contains the suported extensions.
    extensions = ['txt', 'csv', 'shp', 'dbf', 'xlsx', 'parquet']
    try:
        ext = name.split(".")[1].lower()
        assert ext in extensions, f'{ext:} is not supported. Options are {extensions}'
    except IndexError:
        raise Exception('Param "name" was provided without extension')

    if ext in ['shp', 'dbf']:
        df = load_shp_file(path, name, expected_cols)      
        if df.geometry.isnull().all():
            del df['geometry']
    elif ext in ['parquet']:
        df = pd.read_parquet(path + '\\' + name, columns=expected_cols)
    else:        
        df = load_text_file(path, name, expected_cols, usa, **kwargs)

    return df


# SAVE FILES
def save_df_as_csv(df, path, name='test', american=False, kwargs=None):
    """
    Saves a data frame in a CSV
    :param df: dataFrame
    :param path: path to save the file
    :param name: name of the files to save, default as 'test'
    :param american: bool, True or False for american standards
    :param kwargs: args for df.to_csv method from pandas
    :return: nothing
    """
    if kwargs is None:
        kwargs = {}

    keywords = get_american_standers(american)
    kwargs = {**keywords, **kwargs}

    # Reasures passed name is without extension
    file_name = file_name_with_extension(path, name)
    df.to_csv(file_name, **kwargs)
    return True


def save_df_as_parquet(df, path, name):
    try:
        df.columns = df.columns.astype(str)
        file_name = file_name_with_extension(path, name, ext='.parquet')
        df.to_parquet(file_name, compression=None)
    except ValueError:
        print('Error while creating a usable Parquet, creating a CSV instead...')
        save_df_as_csv(df, path, name, american=False)
        
    return True


def save_df_as_shp(df, path, name):
    # There is a limitation with SHPs that only saves columns with 10charcMax.
    # So we are doing a gambiarra...
    
    col_names = list(df.columns)
    col_names.remove('geometry')
    
    aux_dict = dict(zip(col_names, ['c'+str(x) for x in list(range(0, len(col_names)))]))
    temp = pd.Series(aux_dict).to_frame()
    temp.to_csv(os.path.normpath(os.path.join(path, name + '.col')), 
                header=False, index=True)
    
    df = df.rename(columns=aux_dict)    
    df.to_file(os.path.normpath(os.path.join(path, name + '.shp')), 
               driver='ESRI Shapefile')
    return True


def save_df_as_excel(df, path, name, sheet_name='Sheet1', startrow=None,
                     truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [path, name]
    into [sheet_name] Sheet.
    If [path, name] doesn't exist, then this function will create it.

    Parameters:
      df : dataframe to save to workbook
      path:
      name:
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    filename = file_name_with_extension(path, name, ext='.xlsx')
           
    # Ignore [engine] parameter if it was passed.
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # Try to open an existing workbook.
        writer.book = load_workbook(filename)

        # Get the last row in the existing Excel sheet.
        # If it was not specified explicitly.
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # Truncate sheet.
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # Index of [sheet_name] sheet.
            idx = writer.book.sheetnames.index(sheet_name)
            # Remove [sheet_name].
            writer.book.remove(writer.book.worksheets[idx])
            # Create an empty sheet [sheet_name] using old index.
            writer.book.create_sheet(sheet_name, idx)

        # Copy existing sheets.
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, 
                index=False, **to_excel_kwargs)

    # save the workbook
    writer.save()
    return None


# TODO see if this function works and add the option do data.save_file()
def save_zip(df, path, zip_file, file_name):
    # first needs to save the file normally
    temp_file = file_name_with_extension(path, file_name, ext='.txt')
    df.to_csv(temp_file, index=None, sep=';', mode='w', float_format='%.6f',
              decimal=',', date_format='%d/%m/%Y %H:%M:%S')
    # then add its to the zip
    # TODO with file already in zip delete it first
    zip_path = file_name_with_extension(path, zip_file, ext='.zip')
    with ZipFile(zip_path, 'a') as z:
        z.write(temp_file, file_name + '.txt')
    # now delete the temp file
    os.remove(temp_file)

    return True
