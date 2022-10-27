import os
import pandas as pd
import geopandas as gpd

from zipfile import ZipFile
from openpyxl import load_workbook


# HELPERS
def get_american_standers(flag):

    if flag:
        keys = {'sep': ',', 'decimal': '.'}
    else:
        keys = {'sep': ';', 'decimal': ','}
    return keys


def file_name_with_extension(path, name, ext='.csv'):
    # Reassures passed name is without extension.
    name = os.path.splitext(os.path.basename(name))[0]
    name_ext = os.path.normpath(os.path.join(path, name)) + ext
    return name_ext


# LOAD FILES
def load_text_file(path, name, cols=None, usa=False, kwargs=None):

    if kwargs is None:
        kwargs = {}

    keywords = get_american_standers(usa)
    kwargs = {**keywords, **kwargs}
    
    if 'engine' not in kwargs.keys(): 
        if 'low_memory' not in kwargs.keys():
            kwargs['low_memory'] = False
    
    # Escolhe entre ler csv ou excel.
    def read(pathname):
        if name.lower().endswith('.xlsx'):            
            kwargs.pop('low_memory', None)
            kwargs.pop('sep', None)
            kwargs.pop('decimal', None)
            df_ltf = pd.read_excel(pathname, usecols=cols, **kwargs)
        else:
            df_ltf = pd.read_csv(pathname, usecols=cols, **kwargs)
        return df_ltf
    
    # Trata caso esteja ZIPADO.
    if path[-3:] == 'zip':
        with ZipFile(os.path.normpath(path)) as z:
            with z.open(name) as myFile:
                df = read(myFile)                
    else:
        df = read(os.path.normpath(os.path.join(path, name)))

    # Drop rows that contains only NaN
    df.dropna(how="all", inplace=True)
    # Drop cols that contains only NaN
    if df.shape[0] > 1:
        df.dropna(how="all", inplace=True, axis=1)

    return df


def load_shp_file(path, name, cols=None):

    df = gpd.read_file(os.path.normpath(os.path.join(path, name)))
    try:
        # Look for the auxiliary document with the compleat name of columns.
        aux = os.path.join(path, name[:-3]+'col')
        aux = pd.read_csv(aux, header=None, engine='python')
        aux_dict = dict(zip(aux[1], aux[0]))
        df = df.rename(columns=aux_dict)
    except FileNotFoundError:
        pass
    
    if cols is not None:
        for c in cols:
            if c not in df.columns:
                msg_error = "File {} is opened somewhere or missing column {}"\
                    .format(name, c)
                raise Exception(msg_error)
            else:
                pass
        df = df.loc[:, cols]  # Return only expected columns
    return df


def load_file(path, name, expected_cols, usa, **kwargs):
    
    if name is None:
        name = os.path.basename(path)
        path = os.path.dirname(path)
    
    # Checks if name contains the suported extensions.
    extensions = ['txt', 'csv', 'shp', 'dbf', 'xlsx', 'parquet']
    try:
        ext = name.split(".")[1].lower()
        assert ext in extensions, f'{ext:} is not supported. Options are {extensions}'
    except IndexError:
        raise Exception('Param "name" was provided without extension')

    if ext in ['shp', 'dbf']:
        df = load_shp_file(path, name, expected_cols)      
        if df.geometry.isnull().all():
            del df['geometry']
    elif ext in ['parquet']:
        df = pd.read_parquet(path + '\\' + name, columns=expected_cols)
    else:        
        df = load_text_file(path, name, expected_cols, usa, **kwargs)

    return df


# SAVE FILES
def save_df_as_csv(df, path, name='test', american=False, kwargs=None):

    if kwargs is None:
        kwargs = {}

    keywords = get_american_standers(american)
    kwargs = {**keywords, **kwargs}

    # Reasures passed name is without extension
    file_name = file_name_with_extension(path, name)
    df.to_csv(file_name, **kwargs)
    return True


def save_df_as_parquet(df, path, name):
    try:
        df.columns = df.columns.astype(str)
        file_name = file_name_with_extension(path, name, ext='.parquet')
        df.to_parquet(file_name, compression=None)
    except ValueError:
        print('Error while creating a usable Parquet, creating a CSV instead...')
        save_df_as_csv(df, path, name, american=False)
        
    return True


def save_df_as_shp(df, path, name):
    # There is a limitation with SHPs that only saves columns with 10charcMax.
    # So we are doing a gambiarra...
    
    col_names = list(df.columns)
    col_names.remove('geometry')
    
    aux_dict = dict(zip(col_names, ['c'+str(x) for x in list(range(0, len(col_names)))]))
    temp = pd.Series(aux_dict).to_frame()
    temp.to_csv(os.path.normpath(os.path.join(path, name + '.col')), 
                header=False, index=True)
    
    df = df.rename(columns=aux_dict)    
    df.to_file(os.path.normpath(os.path.join(path, name + '.shp')), 
               driver='ESRI Shapefile')
    return True


def save_df_as_excel(df, path, name, sheet_name='Sheet1', startrow=None,
                     truncate_sheet=False, **to_excel_kwargs):

    filename = file_name_with_extension(path, name, ext='.xlsx')
           
    # Ignore [engine] parameter if it was passed.
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    try:
        # Try to open an existing workbook.
        load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Get the last row in the existing Excel sheet.
        # If it was not specified explicitly.
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # Truncate sheet.
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # Index of [sheet_name] sheet.
            idx = writer.book.sheetnames.index(sheet_name)
            # Remove [sheet_name].
            writer.book.remove(writer.book.worksheets[idx])
            # Create an empty sheet [sheet_name] using old index.
            writer.book.create_sheet(sheet_name, idx)

        # Copy existing sheets.
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # File does not exist yet, we will create it.
        pass

    if startrow is None:
        startrow = 0

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Write out the new sheet.
    df.to_excel(writer, sheet_name, startrow=startrow, 
                index=False, **to_excel_kwargs)

    # Save the workbook.
    writer.save()
    return None


def save_zip(df, path, zip_file, file_name):
    # First needs to save the file normally.
    temp_file = file_name_with_extension(path, file_name, ext='.txt')
    df.to_csv(temp_file, index=None, sep=';', mode='w', float_format='%.6f',
              decimal=',', date_format='%d/%m/%Y %H:%M:%S')
    # Then add its to the zip.
    zip_path = file_name_with_extension(path, zip_file, ext='.zip')
    with ZipFile(zip_path, 'a') as z:
        z.write(temp_file, file_name + '.txt')
    # Now delete the temp file.
    os.remove(temp_file)

    return True
